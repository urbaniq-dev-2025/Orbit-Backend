from __future__ import annotations

from io import BytesIO
from uuid import UUID

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from clarivo_ingestion.main import app

client = TestClient(app)


def test_health_endpoint() -> None:
    response = client.get("/healthz")
    assert response.status_code == 200
    assert response.json() == {"status": "ok"}


def test_create_document_with_text_triggers_clarification() -> None:
    payload = {
        "source_type": "pasted_text",
        "content": "short note",
        "metadata": {"client_name": "Acme"},
    }

    response = client.post("/v1/documents", json=payload)
    assert response.status_code == 202
    data = response.json()
    assert UUID(data["doc_id"])
    assert data["status"] in {"awaiting_clarification", "ready_for_preprocessing"}

    status_response = client.get(f"/v1/documents/{data['doc_id']}/status")
    assert status_response.status_code == 200
    status_payload = status_response.json()
    assert status_payload["doc_id"] == data["doc_id"]
    assert status_payload["status"] in {"awaiting_clarification", "ready_for_preprocessing"}
    assert status_payload["scope_available"] is False


def test_submit_clarification_response() -> None:
    payload = {
        "source_type": "pasted_text",
        "content": "short note",
    }
    create_response = client.post("/v1/documents", json=payload)
    doc_id = create_response.json()["doc_id"]

    clarifications = client.get(f"/v1/documents/{doc_id}/clarifications").json()["items"]
    clarification_id = clarifications[0]["clarification_id"]

    response = client.post(
        f"/v1/documents/{doc_id}/clarifications/{clarification_id}/responses",
        json={"answer": "Additional personas include sales and support."},
    )
    assert response.status_code == 200
    assert response.json()["status"] == "answered"

    status_response = client.get(f"/v1/documents/{doc_id}/status")
    assert status_response.json()["status"] == "ready_for_preprocessing"
    assert status_response.json()["scope_available"] is True

    scope_response = client.get(f"/v1/documents/{doc_id}/scope")
    assert scope_response.status_code == 200
    assert scope_response.json()["features"]


def test_cancel_document() -> None:
    payload = {
        "source_type": "pasted_text",
        "content": "another short note",
    }
    create_response = client.post("/v1/documents", json=payload)
    doc_id = create_response.json()["doc_id"]

    cancel_response = client.post(f"/v1/documents/{doc_id}:cancel")
    assert cancel_response.status_code == 200
    assert cancel_response.json()["status"] == "cancelled"


def test_scope_generates_for_long_content() -> None:
    base_sentence = (
        "The system must allow sales managers to manage leads, track deals, integrate with the CRM API, "
        "support real-time dashboards, enforce authentication, and deliver responses under two seconds. "
    )
    content = base_sentence * 8  # ensures length above clarification threshold
    payload = {
        "source_type": "pasted_text",
        "content": content,
    }
    create_response = client.post("/v1/documents", json=payload)
    doc_id = create_response.json()["doc_id"]

    status_response = client.get(f"/v1/documents/{doc_id}/status")
    status_payload = status_response.json()
    assert status_payload["status"] == "ready_for_preprocessing"
    assert status_payload["scope_available"] is True
    assert "scope" in status_payload["links"]

    scope_response = client.get(f"/v1/documents/{doc_id}/scope")
    assert scope_response.status_code == 200
    scope = scope_response.json()
    assert len(scope["features"]) >= 1
    assert len(scope["technical_requirements"]) >= 1

    excel_response = client.get(f"/v1/documents/{doc_id}/scope.xlsx")
    assert excel_response.status_code == 200
    assert excel_response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(excel_response.content))
    assert workbook.sheetnames == ["Features"]
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert headers == [
        "Modules",
        "Features",
        "Interactions",
        "Notes",
        "Questions/Clarifications",
        "Answers",
    ]

    pdf_response = client.get(f"/v1/documents/{doc_id}/scope.pdf")
    assert pdf_response.status_code == 200
    assert pdf_response.headers["content-type"] == "application/pdf"
    assert len(pdf_response.content) > 100

    modules_response = client.get(f"/v1/documents/{doc_id}/modules")
    assert modules_response.status_code == 200
    modules_payload = modules_response.json()
    assert modules_payload["doc_id"] == doc_id
    assert len(modules_payload["modules"]) >= 1
    assert "features" in modules_payload["modules"][0]


def test_scope_endpoint_returns_404_when_missing() -> None:
    payload = {
        "source_type": "pasted_text",
        "content": "short",
    }
    create_response = client.post("/v1/documents", json=payload)
    doc_id = create_response.json()["doc_id"]

    response = client.get(f"/v1/documents/{doc_id}/scope")
    assert response.status_code == 404

    response_excel = client.get(f"/v1/documents/{doc_id}/scope.xlsx")
    assert response_excel.status_code == 404

    response_pdf = client.get(f"/v1/documents/{doc_id}/scope.pdf")
    assert response_pdf.status_code == 404


def test_scope_preview_generates_sections() -> None:
    content = """
    Project Vision: Build a unified CRM dashboard for sales managers.
    Persona: Sales Manager - needs insights into pipeline performance.
    Persona: Customer Success Lead - tracks onboarding tasks.
    Module: Reporting
    - should allow users to filter deals by stage
    - system must integrate with the legacy API for real-time updates.
    Performance should be under 2 seconds per query.
    Are there regional compliance considerations?
    """

    response = client.post("/v1/scope/preview", json={"content": content})
    assert response.status_code == 200
    payload = response.json()
    assert payload["executive_summary"]["overview"]
    assert len(payload["personas"]) >= 1
    assert len(payload["features"]) >= 1
    assert any(
        req["statement"].lower().startswith("system must integrate")
        for req in payload["technical_requirements"]
    )
    assert payload["open_questions"][0]["question"].endswith("?")


def test_scope_preview_requires_content() -> None:
    response = client.post("/v1/scope/preview", json={"content": "   "})
    assert response.status_code == 422

