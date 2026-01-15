"""
Export service for generating Excel and PDF reports from admin dashboard data.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.ext.asyncio import AsyncSession

from app.services import admin as admin_service
from app.models import CreditPurchase, Subscription, Workspace, WorkspaceCreditBalance


async def generate_excel_report(session: AsyncSession) -> bytes:
    """Generate Excel report with admin dashboard data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Admin Dashboard Report"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Title style
    title_font = Font(bold=True, size=16)
    
    # Subtitle style
    subtitle_font = Font(bold=True, size=12, italic=True)
    
    row = 1
    
    # Title
    ws.merge_cells(f"A{row}:D{row}")
    title_cell = ws[f"A{row}"]
    title_cell.value = "Admin Dashboard Report"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    
    # Generated date
    ws.merge_cells(f"A{row}:D{row}")
    date_cell = ws[f"A{row}"]
    date_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    date_cell.font = Font(size=10, italic=True)
    date_cell.alignment = Alignment(horizontal="center")
    row += 2
    
    # 1. Admin Stats
    ws[f"A{row}"] = "1. Platform Statistics"
    ws[f"A{row}"].font = subtitle_font
    row += 1
    
    stats = await admin_service.get_admin_stats(session)
    stats_headers = ["Metric", "Value"]
    ws.append(stats_headers)
    for i, header in enumerate(stats_headers, start=1):
        cell = ws.cell(row=row, column=i)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    stats_data = [
        ["Total Users", stats.get("totalUsers", 0)],
        ["Active Users", stats.get("activeUsers", 0)],
        ["Total Workspaces", stats.get("totalWorkspaces", 0)],
        ["Total Projects", stats.get("totalProjects", 0)],
        ["Total Scopes", stats.get("totalScopes", 0)],
        ["Total Quotations", stats.get("totalQuotations", 0)],
        ["Total Proposals", stats.get("totalProposals", 0)],
        ["Total AI Requests", stats.get("totalAiRequests", 0)],
        ["Total Storage (GB)", stats.get("totalStorageGb", 0)],
    ]
    for stat_row in stats_data:
        ws.append(stat_row)
        row += 1
    
    row += 2
    
    # 2. Revenue Breakdown
    ws[f"A{row}"] = "2. Revenue Breakdown"
    ws[f"A{row}"].font = subtitle_font
    row += 1
    
    revenue_data = await admin_service.get_revenue_breakdown(session)
    
    # Total MRR/ARR
    ws.append(["Total MRR", f"${revenue_data.get('totalMrr', 0):,.2f}"])
    row += 1
    ws.append(["Total ARR", f"${revenue_data.get('totalArr', 0):,.2f}"])
    row += 1
    
    # Revenue by Plan
    ws.append([])
    row += 1
    ws[f"A{row}"] = "Revenue by Plan"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    row += 1
    
    plan_headers = ["Plan", "Count", "Revenue ($)"]
    ws.append(plan_headers)
    for i, header in enumerate(plan_headers, start=1):
        cell = ws.cell(row=row, column=i)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    for plan_info in revenue_data.get("revenueByPlan", []):
        ws.append([
            plan_info.get("plan", "").title(),
            plan_info.get("count", 0),
            f"${plan_info.get('revenue', 0):,.2f}",
        ])
        row += 1
    
    # MRR Breakdown
    ws.append([])
    row += 1
    ws[f"A{row}"] = "MRR Breakdown (Last 6 Months)"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    row += 1
    
    mrr_headers = ["Month", "MRR ($)"]
    ws.append(mrr_headers)
    for i, header in enumerate(mrr_headers, start=1):
        cell = ws.cell(row=row, column=i)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    for month_data in revenue_data.get("mrrBreakdown", []):
        ws.append([
            month_data.get("month", ""),
            f"${month_data.get('mrr', 0):,.2f}",
        ])
        row += 1
    
    row += 2
    
    # 3. Users List (first 50)
    ws[f"A{row}"] = "3. Users List (Top 50)"
    ws[f"A{row}"].font = subtitle_font
    row += 1
    
    users_data = await admin_service.get_users_list(session, page=1, page_size=50)
    user_headers = ["Email", "Full Name", "Active", "Verified", "Onboarding Completed", "Workspaces", "Created At"]
    ws.append(user_headers)
    for i, header in enumerate(user_headers, start=1):
        cell = ws.cell(row=row, column=i)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    for user in users_data.get("users", [])[:50]:
        created_at = user.get("createdAt")
        if created_at:
            if isinstance(created_at, str):
                created_at_str = created_at.split("T")[0]
            else:
                created_at_str = created_at.strftime("%Y-%m-%d")
        else:
            created_at_str = ""
        ws.append([
            user.get("email", ""),
            user.get("fullName", ""),
            "Yes" if user.get("isActive", False) else "No",
            "Yes" if user.get("isVerified", False) else "No",
            "Yes" if user.get("onboardingCompleted", False) else "No",
            user.get("workspaceCount", 0),
            created_at_str,
        ])
        row += 1
    
    row += 2
    
    # 4. Subscriptions List (first 50)
    ws[f"A{row}"] = "4. Subscriptions List (Top 50)"
    ws[f"A{row}"].font = subtitle_font
    row += 1
    
    subscriptions_data = await admin_service.get_subscriptions_list(session, page=1, page_size=50)
    sub_headers = ["Workspace", "Plan", "Status", "Billing Cycle", "Created At"]
    ws.append(sub_headers)
    for i, header in enumerate(sub_headers, start=1):
        cell = ws.cell(row=row, column=i)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    for sub in subscriptions_data.get("subscriptions", [])[:50]:
        created_at = sub.get("createdAt")
        if created_at:
            if isinstance(created_at, str):
                created_at_str = created_at.split("T")[0]
            else:
                created_at_str = created_at.strftime("%Y-%m-%d")
        else:
            created_at_str = ""
        ws.append([
            sub.get("workspaceName", ""),
            sub.get("plan", "").title(),
            sub.get("status", "").title(),
            sub.get("billingCycle", "").title() if sub.get("billingCycle") else "N/A",
            created_at_str,
        ])
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            try:
                # Skip merged cells
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


async def generate_pdf_report(session: AsyncSession) -> bytes:
    """Generate PDF report with admin dashboard data."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = styles["Title"]
    title = Paragraph("Admin Dashboard Report", title_style)
    story.append(title)
    story.append(Spacer(1, 0.2*inch))
    
    # Generated date
    date_style = styles["Normal"]
    date_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    date_para = Paragraph(date_text, date_style)
    story.append(date_para)
    story.append(Spacer(1, 0.3*inch))
    
    # 1. Platform Statistics
    heading_style = styles["Heading2"]
    heading = Paragraph("1. Platform Statistics", heading_style)
    story.append(heading)
    story.append(Spacer(1, 0.2*inch))
    
    stats = await admin_service.get_admin_stats(session)
    stats_data = [
        ["Metric", "Value"],
        ["Total Users", str(stats.get("totalUsers", 0))],
        ["Active Users", str(stats.get("activeUsers", 0))],
        ["Total Workspaces", str(stats.get("totalWorkspaces", 0))],
        ["Total Projects", str(stats.get("totalProjects", 0))],
        ["Total Scopes", str(stats.get("totalScopes", 0))],
        ["Total Quotations", str(stats.get("totalQuotations", 0))],
        ["Total Proposals", str(stats.get("totalProposals", 0))],
        ["Total AI Requests", str(stats.get("totalAiRequests", 0))],
        ["Total Storage (GB)", str(stats.get("totalStorageGb", 0))],
    ]
    
    stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
    stats_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 12),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 0.3*inch))
    
    # 2. Revenue Breakdown
    heading = Paragraph("2. Revenue Breakdown", heading_style)
    story.append(heading)
    story.append(Spacer(1, 0.2*inch))
    
    revenue_data = await admin_service.get_revenue_breakdown(session)
    
    revenue_summary = [
        ["Total MRR", f"${revenue_data.get('totalMrr', 0):,.2f}"],
        ["Total ARR", f"${revenue_data.get('totalArr', 0):,.2f}"],
    ]
    revenue_summary_table = Table(revenue_summary, colWidths=[3*inch, 2*inch])
    revenue_summary_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(revenue_summary_table)
    story.append(Spacer(1, 0.2*inch))
    
    # Revenue by Plan
    subheading = Paragraph("<b>Revenue by Plan</b>", styles["Normal"])
    story.append(subheading)
    story.append(Spacer(1, 0.1*inch))
    
    plan_data = [["Plan", "Count", "Revenue ($)"]]
    for plan_info in revenue_data.get("revenueByPlan", []):
        plan_data.append([
            plan_info.get("plan", "").title(),
            str(plan_info.get("count", 0)),
            f"${plan_info.get('revenue', 0):,.2f}",
        ])
    
    plan_table = Table(plan_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
    plan_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(plan_table)
    story.append(Spacer(1, 0.2*inch))
    
    # MRR Breakdown
    subheading = Paragraph("<b>MRR Breakdown (Last 6 Months)</b>", styles["Normal"])
    story.append(subheading)
    story.append(Spacer(1, 0.1*inch))
    
    mrr_data = [["Month", "MRR ($)"]]
    for month_data in revenue_data.get("mrrBreakdown", []):
        mrr_data.append([
            month_data.get("month", ""),
            f"${month_data.get('mrr', 0):,.2f}",
        ])
    
    mrr_table = Table(mrr_data, colWidths=[2*inch, 3*inch])
    mrr_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(mrr_table)
    story.append(Spacer(1, 0.3*inch))
    
    # 3. Users List (Top 20)
    heading = Paragraph("3. Users List (Top 20)", heading_style)
    story.append(heading)
    story.append(Spacer(1, 0.2*inch))
    
    users_data = await admin_service.get_users_list(session, page=1, page_size=20)
    user_data = [["Email", "Full Name", "Active", "Verified", "Workspaces", "Created"]]
    for user in users_data.get("users", [])[:20]:
        created_at = user.get("createdAt")
        if created_at:
            if isinstance(created_at, str):
                created_at_str = created_at.split("T")[0]
            else:
                created_at_str = created_at.strftime("%Y-%m-%d")
        else:
            created_at_str = ""
        user_data.append([
            user.get("email", "")[:30],  # Truncate long emails
            user.get("fullName", "")[:20] if user.get("fullName") else "N/A",
            "Yes" if user.get("isActive", False) else "No",
            "Yes" if user.get("isVerified", False) else "No",
            str(user.get("workspaceCount", 0)),
            created_at_str,
        ])
    
    user_table = Table(user_data, colWidths=[2*inch, 1.5*inch, 0.7*inch, 0.7*inch, 0.7*inch, 1*inch])
    user_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(user_table)
    story.append(Spacer(1, 0.3*inch))
    
    # 4. Subscriptions List (Top 20)
    heading = Paragraph("4. Subscriptions List (Top 20)", heading_style)
    story.append(heading)
    story.append(Spacer(1, 0.2*inch))
    
    subscriptions_data = await admin_service.get_subscriptions_list(session, page=1, page_size=20)
    sub_data = [["Workspace", "Plan", "Status", "Billing Cycle", "Created"]]
    for sub in subscriptions_data.get("subscriptions", [])[:20]:
        created_at = sub.get("createdAt")
        if created_at:
            if isinstance(created_at, str):
                created_at_str = created_at.split("T")[0]
            else:
                created_at_str = created_at.strftime("%Y-%m-%d")
        else:
            created_at_str = ""
        sub_data.append([
            sub.get("workspaceName", "")[:25],
            sub.get("plan", "").title(),
            sub.get("status", "").title(),
            sub.get("billingCycle", "").title() if sub.get("billingCycle") else "N/A",
            created_at_str,
        ])
    
    sub_table = Table(sub_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1*inch])
    sub_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(sub_table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


async def generate_subscriptions_export(
    session: AsyncSession,
    format: str = "xlsx",
    status: Optional[str] = None,
    plan: Optional[str] = None,
    search: Optional[str] = None,
) -> bytes:
    """Generate subscriptions export in Excel or CSV format."""
    # Get subscription list data
    list_data = await admin_service.get_subscription_list_enhanced(
        session,
        page=1,
        page_size=10000,  # Get all for export
        search=search,
        status=status,
        plan=plan,
    )
    
    if format == "csv":
        # Generate CSV
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow([
            "Customer Name",
            "Email",
            "Plan",
            "Status",
            "MRR",
            "Credits",
            "Started Date",
            "Renewal Date",
            "Billing Cycle",
        ])
        
        # Data rows
        for sub in list_data.get("subscriptions", []):
            started = sub.get("started")
            started_str = started.strftime("%Y-%m-%d") if started and hasattr(started, "strftime") else (started.split("T")[0] if isinstance(started, str) else "")
            renews = sub.get("renews")
            renews_str = renews.strftime("%Y-%m-%d") if renews and hasattr(renews, "strftime") else (renews.split("T")[0] if isinstance(renews, str) else "") if renews else ""
            
            writer.writerow([
                sub.get("customer", ""),
                sub.get("email", ""),
                sub.get("plan", ""),
                sub.get("status", ""),
                sub.get("mrr", 0),
                sub.get("credits", 0),
                started_str,
                renews_str,
                sub.get("billingCycle", "") or "",
            ])
        
        return output.getvalue().encode("utf-8")
    else:
        # Generate Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Subscriptions Export"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Headers
        headers = [
            "Customer Name",
            "Email",
            "Plan",
            "Status",
            "MRR",
            "Credits",
            "Started Date",
            "Renewal Date",
            "Billing Cycle",
        ]
        ws.append(headers)
        
        # Style headers
        for i, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=i)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for sub in list_data.get("subscriptions", []):
            started = sub.get("started")
            started_str = started.strftime("%Y-%m-%d") if started and hasattr(started, "strftime") else (started.split("T")[0] if isinstance(started, str) else "")
            renews = sub.get("renews")
            renews_str = renews.strftime("%Y-%m-%d") if renews and hasattr(renews, "strftime") else (renews.split("T")[0] if isinstance(renews, str) else "") if renews else ""
            
            ws.append([
                sub.get("customer", ""),
                sub.get("email", ""),
                sub.get("plan", ""),
                sub.get("status", ""),
                sub.get("mrr", 0),
                sub.get("credits", 0),
                started_str,
                renews_str,
                sub.get("billingCycle", "") or "",
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, "column_letter"):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


async def generate_credit_purchases_export(
    session: AsyncSession,
    format: str = "xlsx",
    package: Optional[str] = None,
    search: Optional[str] = None,
) -> bytes:
    """Generate credit purchases export in Excel or CSV format."""
    # Get credit purchases data
    purchases_data = await admin_service.get_credit_purchases(
        session,
        page=1,
        page_size=10000,  # Get all for export
        search=search,
        package=package,
    )
    
    if format == "csv":
        # Generate CSV
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow([
            "Customer Name",
            "Package",
            "Amount",
            "Credits",
            "Date",
            "Payment Method",
            "Transaction ID",
            "Status",
        ])
        
        # Data rows
        for purchase in purchases_data.get("purchases", []):
            date = purchase.get("date")
            date_str = date.strftime("%Y-%m-%d") if date and hasattr(date, "strftime") else (date.split("T")[0] if isinstance(date, str) else "")
            
            writer.writerow([
                purchase.get("customer", ""),
                purchase.get("package", ""),
                purchase.get("amount", 0),
                purchase.get("credits", 0),
                date_str,
                purchase.get("method", ""),
                purchase.get("transactionId", ""),
                purchase.get("status", ""),
            ])
        
        return output.getvalue().encode("utf-8")
    else:
        # Generate Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Credit Purchases Export"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Headers
        headers = [
            "Customer Name",
            "Package",
            "Amount",
            "Credits",
            "Date",
            "Payment Method",
            "Transaction ID",
            "Status",
        ]
        ws.append(headers)
        
        # Style headers
        for i, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=i)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for purchase in purchases_data.get("purchases", []):
            date = purchase.get("date")
            date_str = date.strftime("%Y-%m-%d") if date and hasattr(date, "strftime") else (date.split("T")[0] if isinstance(date, str) else "")
            
            ws.append([
                purchase.get("customer", ""),
                purchase.get("package", ""),
                purchase.get("amount", 0),
                purchase.get("credits", 0),
                date_str,
                purchase.get("method", ""),
                purchase.get("transactionId", ""),
                purchase.get("status", ""),
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, "column_letter"):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
